from robobrowser import RoboBrowser
from time import sleep
import openpyxl
from json import dumps
from requests_html import HTMLSession
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
import pickle
import selenium.webdriver

def Login_and_get_get_Cookeis():
    browser = RoboBrowser(history=True , parser="lxml")
    browser.open('http://smc.smcegy.com/smc/Home/Index')
    form = browser.get_form(action='/smc/Home/Index')
    form["username"].value = "اسماء لطفي76"
    form["password"].value = "456789"
    browser.submit_form(form)
    return browser.session.cookies.get_dict()

cookies = Login_and_get_get_Cookeis()


browser = webdriver.Chrome()
browser.get('http://smc.smcegy.com/smc/Home/Index')



# #Load the Excel file
# excelpath = "test.xlsx";wb_obj = openpyxl.load_workbook(excelpath)
# sheet_obj = wb_obj.active;cell_obj = sheet_obj.cell(row=7, column=3);start = 8
# counter = start;currentpat = sheet_obj.cell(row=counter, column=3)

# while currentpat.value is not None:
#     currentpat = sheet_obj.cell(row=counter, column=3)
#     link = "http://smc.smcegy.com/smc/HospDecreeReceipts/Create?decreeId="
#     PatLink = link+ str(currentpat.value)
#     print(PatLink)
#     browser.get(PatLink)
#     try :
#         chaker = browser.find_element_by_id("#addBtn")
#         if chaker :
#             print("found -------------------")
        
#     except :print("not find ")
#     sleep(10)
 
    
#     counter = counter + 1
#     currentpat = sheet_obj.cell(row=counter, column=3)




















#    Dom = HTMLSession()
#    get_site = Dom.get(PatLink, cookies=cookies)
#     get_sitehtml.render("")
#     with open("file.txt" , "a") as html : html.write(res.html.text)